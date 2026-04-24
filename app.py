import streamlit as st
import pandas as pd
from io import BytesIO
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from streamlit_gsheets import GSheetsConnection

# --- 1. ตั้งค่าพื้นฐานและเชื่อมต่อ GSheets ---
st.set_page_config(page_title="วิทยาลัยเทคโนโลยีนนทบุรี", layout="wide")
conn = st.connection("gsheets", type=GSheetsConnection)

def load_data():
    try:
        data = conn.read(spreadsheet=st.secrets["gsheet_url"], ttl=0)
        for col in data.columns:
            data[col] = data[col].astype(str).str.replace(r'\.0$', '', regex=True)
            data[col] = data[col].str.replace("'", "").replace('nan', '')
        return data
    except:
        return pd.DataFrame(columns=['รุ่น', 'รหัสนักศึกษา', 'ชื่อ', 'นามสกุล', 'ระดับชั้น', 'Room'])

def get_logo_image():
    # ตรวจสอบว่ามีไฟล์โลโก้อยู่หรือไม่
    if os.path.exists("logo_college.jpg"):
        try:
            return XLImage("logo_college.jpg")
        except: return None
    return None

# --- 2. ฟังก์ชันสร้างใบเช็คชื่อ (แบบที่ 1) ---
def create_attendance_report(target_year):
    df_all = load_data()
    if df_all.empty: return None
    year_data = df_all[df_all['ระดับชั้น'] == target_year]
    if year_data.empty: return None
    output = BytesIO()
    wb = Workbook(); wb.remove(wb.active)
    side = Side(style='thin'); border = Border(left=side, right=side, top=side, bottom=side)
    f_bold = Font(name='Angsana New', size=15, bold=True); center = Alignment(horizontal='center', vertical='center')

    for r_name in sorted(year_data['Room'].unique()):
        ws = wb.create_sheet(title=f"เช็คชื่อ-{r_name.replace('/', '-')}")
        room_data = year_data[year_data['Room'] == r_name].sort_values('รหัสนักศึกษา')
        ws.merge_cells('O2:V2'); ws['O2'] = "บัญชีรายชื่อนี้ใช้สำหรับ"; ws['O2'].font = f_bold; ws['O2'].alignment = center; ws['O2'].border = border
        ws.merge_cells('O3:P4'); ws['O3'] = "เช็คชื่อนักศึกษา"; ws['O3'].border = border; ws['O3'].alignment = center
        ws.merge_cells('Q3:S4'); ws['Q3'] = "เซ็นสอบกลางภาค"; ws['Q3'].border = border; ws['Q3'].alignment = center
        ws.merge_cells('T3:V4'); ws['T3'] = "เซ็นสอบปลายภาค"; ws['T3'].border = border; ws['T3'].alignment = center
        ws.merge_cells('A5:V5'); ws['A5'] = "บัญชีรายชื่อนักศึกษา ภาคเรียนที่ 1 ปีการศึกษา 2568"; ws['A5'].font = f_bold; ws['A5'].alignment = center
        ws.merge_cells('A6:V6'); ws['A6'] = f"ระดับ ปวส. ชั้นปีที่ {target_year[2:]} ห้อง {r_name} ศูนย์บางแค"; ws['A6'].font = f_bold; ws['A6'].alignment = center
        ws.merge_cells('A8:A10'); ws['A8'] = "เลขที่"; ws.merge_cells('B8:B10'); ws['B8'] = "รหัสประจำตัว"; ws.merge_cells('C8:D10'); ws['C8'] = "ชื่อ-สกุล"
        ws['E8'] = "เดือน"; ws['E9'] = "วันที่"; ws['E10'] = "คาบ"; ws.merge_cells('V8:V10'); ws['V8'] = "หมายเหตุ"
        for i in range(1, 17): ws.cell(row=10, column=5+i).value = i
        for r in range(8, 11):
            for c in range(1, 23):
                cell = ws.cell(row=r, column=c); cell.border = border; cell.alignment = center; cell.font = f_bold
        for i, row in enumerate(room_data.itertuples(), 1):
            curr = 10 + i
            ws.cell(row=curr, column=1).value = i; ws.cell(row=curr, column=2).value = row.รหัสนักศึกษา
            ws.merge_cells(start_row=curr, start_column=3, end_row=curr, end_column=4)
            ws.cell(row=curr, column=3).value = f"{row.ชื่อ} {row.นามสกุล}"
            for c in range(1, 23): ws.cell(row=curr, column=c).border = border; ws.cell(row=curr, column=c).alignment = center
        ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 15; ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 14
        for c_idx in range(5, 22): ws.column_dimensions[get_column_letter(c_idx)].width = 3.5
        ws.column_dimensions['V'].width = 12
        img = get_logo_image()
        if img: img.width, img.height = 75, 75; ws.add_image(img, 'H1')
    wb.save(output); return output.getvalue()

# --- 3. ฟังก์ชันสร้างฟอร์มเกรด (แบบที่ 2 - ตามสั่งเป๊ะ + มีโลโก้) ---
def create_grade_report(target_year):
    df_all = load_data()
    if df_all.empty: return None
    year_data = df_all[df_all['ระดับชั้น'] == target_year]
    if year_data.empty: return None

    output = BytesIO()
    wb = Workbook(); wb.remove(wb.active)
    side = Side(style='thin'); border = Border(left=side, right=side, top=side, bottom=side)
    f_bold = Font(name='Angsana New', size=14, bold=True)
    f_normal = Font(name='Angsana New', size=13)
    
    # การตั้งค่าแนวอักษร
    rotate_align = Alignment(horizontal='center', vertical='center', textRotation=90)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r_name in sorted(year_data['Room'].unique()):
        ws = wb.create_sheet(title=f"เกรด-{r_name.replace('/', '-')}")
        room_data = year_data[year_data['Room'] == r_name].sort_values('รหัสนักศึกษา')
        
        # แทรกโลโก้ที่หัวกระดาษบัญชีเกรด
        img = get_logo_image()
        if img:
            img.width, img.height = 70, 70
            ws.add_image(img, 'A1') # วางโลโก้ที่มุมซ้ายบน

        # หัวกระดาษ
        ws.merge_cells('A4:R4'); ws['A4'] = "บัญชีผลการเรียนรายวิชา"; ws['A4'].alignment = center_align; ws['A4'].font = f_bold
        ws.merge_cells('A5:R5'); ws['A5'] = "ภาคเรียนที่  ...............  ปีการศึกษา .........................."; ws['A5'].alignment = center_align; ws['A5'].font = f_normal
        ws.merge_cells('A6:C6'); ws['A6'] = "รหัสวิชา  ………………………….."
        ws.merge_cells('D6:L6'); ws['D6'] = "ชื่อวิชา  ……………………………………………………………………………………"
        ws.merge_cells('M6:R6'); ws['M6'] = "หน่วยกิต ……. หน่วยกิต"
        ws.merge_cells('A7:H7'); ws['A7'] = f"ระดับ ปวส. ชั้นปีที่ {target_year[2:]} ห้อง {r_name}"
        ws.merge_cells('I7:R7'); ws['I7'] = "ผู้สอน  ..........................................................................................."
        for cell in ['A6','D6','M6','A7','I7']: ws[cell].font = f_normal

        # --- โครงสร้างหัวตาราง (แถว 8-11) ---
        for col, val in [('A8', 'เลขที่'), ('B8', 'รหัสประจำตัว'), ('C8', 'ชื่อ - สกุล')]:
            ws.merge_cells(f'{col}:{col[0]}11'); ws[col] = val; ws[col].alignment = center_align

        # ทฤษฎี (E-J)
        ws.merge_cells('E8:J8'); ws['E8'] = "ทฤษฎี..........................หน่วยกิต"
        ws.merge_cells('E9:I9'); ws['E9'] = "คะแนนระหว่างภาค"
        headers_theory = {'E10':"เวลา/อุปกรณ์",'F10':"พฤติกรรม",'G10':"งาน/ทดสอบ",'H10':"สอบกลางภาค",'I10':"สอบปลายภาค",'J10':"คะแนนรวม"}
        for cell, val in headers_theory.items(): ws[cell] = val; ws[cell].alignment = rotate_align
        ws.merge_cells('J9:J10')

        # ปฏิบัติ (K-P)
        ws.merge_cells('K8:P8'); ws['K8'] = "ปฏิบัติ..................หน่วยกิต"
        ws.merge_cells('K9:O9'); ws['K9'] = "คะแนนระหว่างภาค"
        headers_practice = {'K10':"คุณภาพของงาน",'L10':"เวลา/อุปกรณ์",'M10':"พฤติกรรม",'N10':"การปฎิบัติงาน",'O10':"สอบทฤษฎีเชิงปฎิบัติ",'P10':"คะแนนรวม"}
        for cell, val in headers_practice.items(): ws[cell] = val; ws[cell].alignment = rotate_align
        ws.merge_cells('P9:P10')

        # ระดับ/หมายเหตุ
        ws.merge_cells('Q8:Q9'); ws['Q8'] = "ระดับ"; ws['Q10'] = "คะแนน"; ws.merge_cells('Q10:Q11')
        ws.merge_cells('R8:R11'); ws['R8'] = "หมายเหตุ"

        # คะแนนเต็มแถว 11
        pts = {5:'10',6:'10',7:'20',8:'20',9:'40',10:'100',11:'20',12:'10',13:'10',14:'40',15:'20',16:'100'}
        for c, v in pts.items(): ws.cell(row=11, column=c).value = v; ws.cell(row=11, column=c).alignment = center_align

        # ตีกรอบหัวตาราง
        for r in range(8, 12):
            for c in range(1, 19):
                cell = ws.cell(row=r, column=c); cell.border = border; cell.font = f_bold
                if not cell.alignment: cell.alignment = center_align

        # --- ตั้งค่าความกว้างคอลัมน์ (ตามสเปกเป๊ะ) ---
        ws.column_dimensions['A'].width = 3.86
        ws.column_dimensions['B'].width = 12.29
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 17
        for c_idx in ['E','F','G','H','I','J','L','M','N','O']: ws.column_dimensions[c_idx].width = 2.86
        ws.column_dimensions['K'].width = 3.29
        ws.column_dimensions['P'].width = 3.29
        ws.column_dimensions['Q'].width = 5.86
        ws.column_dimensions['R'].width = 8

        # --- ข้อมูลนักศึกษา ---
        for i, row in enumerate(room_data.itertuples(), 1):
            curr = 11 + i
            ws.cell(row=curr, column=1).value = i; ws.cell(row=curr, column=2).value = row.รหัสนักศึกษา
            ws.merge_cells(start_row=curr, start_column=3, end_row=curr, end_column=4)
            ws.cell(row=curr, column=3).value = f"{row.ชื่อ} {row.นามสกุล}"
            for c in range(1, 19):
                ws.cell(row=curr, column=c).border = border; ws.cell(row=curr, column=c).alignment = center_align if c != 3 else Alignment(horizontal='left', indent=1)

    wb.save(output); return output.getvalue()

# --- 4. หน้าจอ Streamlit ---
st.title("🏫 ระบบจัดการข้อมูลวิทยาลัย (เวอร์ชันสมบูรณ์ + โลโก้)")

t1, t2, t3 = st.tabs(["📝 ลงทะเบียน", "🔍 แก้ไขข้อมูล", "📥 ดาวน์โหลดเอกสาร"])

# (ส่วน Tab 1 และ 2 เหมือนเดิมเพื่อประหยัดพื้นที่แสดงผล)
with t1:
    with st.form("reg"):
        c1, c2, c3 = st.columns(3)
        with c1: batch = st.text_input("รุ่น"); sid = st.text_input("รหัสนักศึกษา")
        with c2: fname = st.text_input("ชื่อ"); lname = st.text_input("นามสกุล")
        with c3:
            level = st.selectbox("ระดับชั้น", ["ปี1", "ปี2"])
            room = st.selectbox("ห้อง", [f"{'O1' if level=='ปี1' else 'O2'}/{i}" for i in range(1, 16)])
        if st.form_submit_button("💾 บันทึก"):
            if sid and fname:
                df = load_data()
                new = pd.DataFrame([{"รุ่น":f"'{batch}","รหัสนักศึกษา":f"'{sid}","ชื่อ":fname.strip(),"นามสกุล":lname.strip(),"ระดับชั้น":level,"Room":room}])
                conn.update(spreadsheet=st.secrets["gsheet_url"], data=pd.concat([df, new], ignore_index=True))
                st.success("บันทึกเรียบร้อย!"); st.rerun()

with t2:
    df_edit = load_data()
    if not df_edit.empty:
        edited = st.data_editor(df_edit, num_rows="dynamic", use_container_width=True)
        if st.button("💾 บันทึกการแก้ไข"):
            conn.update(spreadsheet=st.secrets["gsheet_url"], data=edited)
            st.success("อัปเดตเรียบร้อย!"); st.rerun()

with t3:
    st.subheader("📥 ดาวน์โหลดเอกสาร (พร้อมโลโก้และฟอร์แมตแนวตั้ง)")
    st.write("---")
    col1, col2 = st.columns(2)
    with col1:
        st.write("**1. ใบเช็คชื่อ (C-V)**")
        f1 = create_attendance_report("ปี1")
        if f1: st.download_button("📥 เช็คชื่อ ปี 1", f1, "Attendance_P1.xlsx")
        f2 = create_attendance_report("ปี2")
        if f2: st.download_button("📥 เช็คชื่อ ปี 2", f2, "Attendance_P2.xlsx")
    
    with col2:
        st.write("**2. บัญชีผลการเรียน (แนวตั้ง)**")
        g1 = create_grade_report("ปี1")
        if g1: st.download_button("📊 ฟอร์มเกรด ปี 1", g1, "Grade_P1.xlsx", type="primary")
        g2 = create_grade_report("ปี2")
        if g2: st.download_button("📊 ฟอร์มเกรด ปี 2", g2, "Grade_P2.xlsx", type="primary")
